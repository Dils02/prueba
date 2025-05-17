import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl import load_workbook
from datetime import datetime
from calculo_mensual.calculo_mensual import procesar_calculo_mensual
from calculo_mensual.rellenar_resumen_anual import rellenar_resumen_anual_totales
from calculo_mensual.recalculo import recalcular_excel_desde_workbook

def mostrar_interfaz_calculo_mensual():
    st.header("📅 Cálculo Mensual")

    archivo = st.file_uploader("📂 Sube el archivo Excel", type=["xlsx"])

    if archivo:
        try:
            wb = load_workbook(archivo)

            meses_validos = [
                "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
                "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"
            ]
            hojas_disponibles = [hoja for hoja in wb.sheetnames if hoja in meses_validos]

            if hojas_disponibles:
                with st.expander("🧮 Calcular Meses Seleccionados"):
                    st.markdown("### ✅ Selecciona los meses a procesar:")

                    seleccionar_todos = st.checkbox("Seleccionar todos los meses")
                    seleccionados = []

                    col1, col2 = st.columns(2)
                    for i, mes in enumerate(meses_validos):
                        if mes in hojas_disponibles:
                            with col1 if i < 6 else col2:
                                estado = seleccionar_todos or st.checkbox(mes, key=f"mes_{mes}")
                                if estado:
                                    seleccionados.append(mes)

                    if st.button("📊 Ejecutar cálculo"):
                        if seleccionados:
                            for mes_seleccionado in seleccionados:
                                ws = wb[mes_seleccionado]
                                data = [[cell.value for cell in row] for row in ws.iter_rows(min_row=2)]
                                columnas = [cell.value for cell in ws[1]]

                                df_mes = pd.DataFrame(data, columns=columnas)
                                df_resultado = procesar_calculo_mensual(df_mes, mes_seleccionado)

                                cols_numericas = df_resultado.select_dtypes(include=["number"]).columns
                                df_resultado[cols_numericas] = df_resultado[cols_numericas].fillna(0)
                                df_resultado = df_resultado.infer_objects(copy=False)

                                for i, row in df_resultado.iterrows():
                                    for j, value in enumerate(row):
                                        ws.cell(row=i + 2, column=j + 1, value=value)

                                rellenar_resumen_anual_totales(wb, mes_seleccionado)

                            excel_file = BytesIO()
                            wb.save(excel_file)
                            excel_file.seek(0)

                            fecha_actual = datetime.now().strftime("%Y-%m-%d")
                            nuevo_nombre = f"PPTO_{fecha_actual}.xlsx"

                            st.success("✅ Archivo procesado exitosamente.")
                            st.download_button(
                                "📥 Descargar archivo procesado",
                                data=excel_file,
                                file_name=nuevo_nombre,
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )
                        else:
                            st.warning("⚠️ Debes seleccionar al menos un mes.")

                with st.expander("🔁 Recalcular Excel"):
                    if st.button("♻️ Ejecutar recalculo"):
                        try:
                            recalcular_excel_desde_workbook(wb)

                            recalculado_file = BytesIO()
                            wb.save(recalculado_file)
                            recalculado_file.seek(0)

                            st.success("✅ Recalculo completado con éxito.")
                            st.download_button(
                                label="📥 Descargar archivo recalculado",
                                data=recalculado_file,
                                file_name="PPTO_Recalculado.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )
                        except Exception as e:
                            st.error(f"❌ Error al recalcular: {e}")
            else:
                st.warning("⚠️ El archivo no contiene hojas con nombres de meses válidos.")
        except Exception as e:
            st.error(f"❌ Ocurrió un error al procesar el archivo: {e}")
