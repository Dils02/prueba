import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side
from openpyxl.utils import get_column_letter

def crear_estructura_resumen_anual(ruta_archivo: str) -> pd.DataFrame:
    df = pd.read_excel(ruta_archivo, sheet_name="Datos")
    cc_col = "Trabajo - Centro de Costo"

    if cc_col not in df.columns:
        raise ValueError("⚠️ Faltan columnas necesarias en la hoja 'Datos'.")

    df[cc_col] = pd.to_numeric(df[cc_col], errors='coerce')
    centros_costo = df[cc_col].dropna().unique()
    centros_costo_filtrados = sorted({int(cc) for cc in centros_costo if 201 <= cc <= 501})

    meses = [
        'Enero', 'Febrero', 'Marzo', 'Abril',
        'Mayo', 'Junio', 'Julio', 'Agosto',
        'Septiembre', 'Octubre', 'Noviembre', 'Diciembre'
    ]

    data = {'Resumen': centros_costo_filtrados}
    for mes in meses:
        data[mes] = ['' for _ in centros_costo_filtrados]
    data['Cantidad Personas C.C'] = [0 for _ in centros_costo_filtrados]

    return pd.DataFrame(data)

def rellenar_hoja_resumen_anual(ruta_archivo):
    wb = load_workbook(ruta_archivo)

    if "Resumen Anual" not in wb.sheetnames:
        raise ValueError("⚠️ La hoja 'Resumen Anual' no existe en el archivo.")

    ws_resumen = wb["Resumen Anual"]

    df = pd.read_excel(ruta_archivo, sheet_name="Datos")
    cc_col = "Trabajo - Centro de Costo"

    if cc_col not in df.columns:
        raise ValueError("⚠️ Faltan columnas necesarias en la hoja 'Datos'.")

    df[cc_col] = pd.to_numeric(df[cc_col], errors='coerce')
    centros_costo = df[cc_col].dropna().unique()
    centros_costo_filtrados = sorted({int(cc) for cc in centros_costo if 201 <= cc <= 501})

    # Contar personas por Centro de Costo
    cantidad_personas = df[cc_col].value_counts().to_dict()

    meses = [
        'Enero', 'Febrero', 'Marzo', 'Abril',
        'Mayo', 'Junio', 'Julio', 'Agosto',
        'Septiembre', 'Octubre', 'Noviembre', 'Diciembre'
    ]

    encabezados = ['Resumen'] + meses + ['Cantidad Personas C.C']

    border_style = Border(
        left=Side(style='medium', color='000000'),
        right=Side(style='medium', color='000000'),
        top=Side(style='medium', color='000000'),
        bottom=Side(style='medium', color='000000')
    )
    font_negrita = Font(bold=True, color="000000")

    # Escribir encabezados con estilo
    for col_idx, encabezado in enumerate(encabezados, start=1):
        cell = ws_resumen.cell(row=1, column=col_idx, value=encabezado)
        cell.font = font_negrita
        cell.border = border_style

    # Escribir datos y aplicar bordes
    for row_idx, centro_costo in enumerate(centros_costo_filtrados, start=2):
        ws_resumen.cell(row=row_idx, column=1, value=centro_costo).border = border_style

        # Inicializar meses en blanco
        for col_idx in range(2, 14):  # columnas 2 a 13 (Enero a Diciembre)
            ws_resumen.cell(row=row_idx, column=col_idx, value='').border = border_style

        # Escribir cantidad de personas por C.C en la columna 14
        cantidad = cantidad_personas.get(centro_costo, 0)
        ws_resumen.cell(row=row_idx, column=14, value=cantidad).border = border_style

    # Ajustar ancho automático de columnas
    for col_idx, encabezado in enumerate(encabezados, start=1):
        max_length = len(str(encabezado))
        for row in range(2, len(centros_costo_filtrados) + 2):
            cell_value = ws_resumen.cell(row=row, column=col_idx).value
            if cell_value is not None:
                max_length = max(max_length, len(str(cell_value)))
        ws_resumen.column_dimensions[get_column_letter(col_idx)].width = max_length + 2

    wb.save(ruta_archivo)
