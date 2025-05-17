import pandas as pd
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl import load_workbook

MESES = [
    "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
    "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"
]

def rellenar_resumen_anual_totales(wb: Workbook, mes_a_actualizar: str):
    try:
        if mes_a_actualizar not in MESES:
            raise ValueError(f"El mes '{mes_a_actualizar}' no es válido.")

        if "Resumen Anual" not in wb.sheetnames:
            print("⚠️ La hoja 'Resumen Anual' no existe en el archivo.")
            return

        if mes_a_actualizar not in wb.sheetnames:
            print(f"⚠️ La hoja '{mes_a_actualizar}' no existe en el archivo.")
            return

        ws_mes: Worksheet = wb[mes_a_actualizar]
        ws_resumen: Worksheet = wb["Resumen Anual"]

        # Obtener encabezados
        headers = [cell.value for cell in ws_mes[1]]
        data_rows = [[cell.value for cell in row] for row in ws_mes.iter_rows(min_row=2)]

        df_mes = pd.DataFrame(data_rows, columns=headers)

        if "Trabajo - Centro de Costo" not in df_mes.columns or "Total Costo Tulsa" not in df_mes.columns:
            print(f"⚠️ La hoja '{mes_a_actualizar}' no tiene las columnas necesarias.")
            return

        resumen_mes = (
            df_mes.groupby("Trabajo - Centro de Costo")["Total Costo Tulsa"]
            .sum()
            .reset_index()
        )

        cc_to_row = {
            ws_resumen.cell(row=i, column=1).value: i
            for i in range(2, ws_resumen.max_row + 1)
        }

        columna_mes = MESES.index(mes_a_actualizar) + 2  # col 1 = centro costo, mes desde col 2

        actualizados = []

        for _, fila in resumen_mes.iterrows():
            cc = fila["Trabajo - Centro de Costo"]
            total = fila["Total Costo Tulsa"]

            if cc in cc_to_row:
                row = cc_to_row[cc]
                ws_resumen.cell(row=row, column=columna_mes).value = total
                actualizados.append((cc, total))

        print(f"✅ Resumen Anual actualizado correctamente para el mes '{mes_a_actualizar}'.")
        if actualizados:
            print("📋 Centros de costo actualizados:")
            for cc, total in actualizados:
                print(f" - {cc}: {total:,.2f}")
        else:
            print("⚠️ No se actualizó ningún centro de costo.")

    except Exception as e:
        print(f"❌ Error al rellenar hoja 'Resumen Anual': {e}")
