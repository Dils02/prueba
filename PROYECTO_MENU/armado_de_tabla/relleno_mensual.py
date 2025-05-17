import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

MESES = [
    "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
    "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"
]

COLUMNAS_REQUERIDAS = [
    "Empleado - Número de Documento",
    "Empleado - Nombre Completo",
    "Trabajo - Centro de Costo",
    "Trabajo - Cargo",
    "Campos Personalizados de Trabajo - Jornada",
    "Trabajo - Sueldo Base"
]

# Nuevas columnas que posiblemente tengan sí/no en Datos
COLUMNAS_SINO = [
    # Lista de columnas en Datos que contienen si/no y deben replicarse
    "Campos Personalizados de Empleado - BAP con descuento",
    "Campos Personalizados de Empleado - Gratificacion",
    "Campos Personalizados de Empleado - Asimilado Sindicato",
    "Campos Personalizados de Empleado - Bono Responsabilidad",
    "Campos Personalizados de Empleado - Bono Operacional",
    "Campos Personalizados de Empleado - Bono asistencia y movilidad",
    "Campos Personalizados de Empleado - Bono Asistencia",
    "Campos Personalizados de Empleado - Bono Turno",
    "Campos Personalizados de Empleado - Monto Asignación Colación",
    "Campos Personalizados de Empleado - Monto Asignación Movili",
    "Campos Personalizados de Empleado - Estudio y desempeño",
    "Campos Personalizados de Empleado - Bono Terreno",
    "Campos Personalizados de Empleado - Bono Tope Seguridad",
    "Campos Personalizados de Empleado - Bono producción Planta",
    "Campos Personalizados de Empleado - Bono producción Nuevo",
    "Campos Personalizados de Empleado - Bono Tope Supervisor"
]

# Columnas extra que no provienen de Datos, inicializadas vacías
COLUMNAS_EXTRA = [
"HE 55% 6x1", "Monto HE 55% 6x1", "HE 100% 6x1", "Monto HE 100% 6x1", "HE 150% 6x1", "Monto HE 150% 6x1",
"HE 50% 6x2", "Monto HE 50% 6x2", "HE 80% 6x2", "Monto HE 80% 6x2", "HE 100% 6x2", "Monto HE 100% 6x2",
"HE 120% 6x2", "Monto HE 120% 6x2", "HE 50% 5x2" , "Monto HE 50% 5x2",
"Total Costo Tulsa"
]

# Columnas de salida totales
COLUMNAS_SALIDA = COLUMNAS_REQUERIDAS + COLUMNAS_SINO + COLUMNAS_EXTRA

def rellenar_hojas_mensuales(ruta_excel):
    try:
        # Leemos la hoja Datos completa
        df_datos = pd.read_excel(ruta_excel, sheet_name="Datos")
        # Preparamos df_filtrado con columnas requeridas
        df_filtrado = df_datos[COLUMNAS_REQUERIDAS].copy()
        # Añadimos las columnas de si/no: si en Datos hay valor, lo copiamos, si no, lo dejamos vacío
        for col in COLUMNAS_SINO:
            if col in df_datos.columns:
                df_filtrado[col] = df_datos[col].fillna("")
            else:
                df_filtrado[col] = ""
        # Añadimos columnas extra vacías
        for col in COLUMNAS_EXTRA:
            df_filtrado[col] = ""

        # Abrimos el libro
        wb = load_workbook(ruta_excel)
        for mes in MESES:
            # Creamos o limpiamos la hoja mensual
            if mes not in wb.sheetnames:
                ws = wb.create_sheet(title=mes)
            else:
                ws = wb[mes]
                ws.delete_rows(1, ws.max_row)

            # Escribir encabezados
            for idx, name in enumerate(COLUMNAS_SALIDA, start=1):
                cell = ws.cell(row=1, column=idx, value=name)
                cell.style = 'Normal'

            # Escribir filas: replicamos mismo valor de df_filtrado
            for row_idx, row in enumerate(df_filtrado.itertuples(index=False), start=2):
                for col_idx, value in enumerate(row, start=1):
                    ws.cell(row=row_idx, column=col_idx, value=value)

            # Ajustar anchos
            for col_idx, name in enumerate(COLUMNAS_SALIDA, start=1):
                col_letter = get_column_letter(col_idx)
                max_len = max(
                    [len(str(cell.value)) for cell in ws[col_letter] if cell.value] + [len(name)]
                )
                ws.column_dimensions[col_letter].width = max_len + 2

        wb.save(ruta_excel)
        print("✅ Hojas mensuales rellenadas correctamente, copiando si/no de Datos.")
    except Exception as e:
        print(f"⚠️ Error al rellenar hojas mensuales: {e}")
