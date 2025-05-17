import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from armado_de_tabla.relleno_HE import rellenar_hoja_horas_extras
from armado_de_tabla.relleno_mensual import rellenar_hojas_mensuales
from armado_de_tabla.relleno_ResumenAnual import rellenar_hoja_resumen_anual

# Función para crear la estructura básica del archivo Excel
def estructurar_hojas_excel(ruta_archivo):
    try:
        wb = load_workbook(ruta_archivo)

        # Renombrar hoja si existe con nombre alternativo
        if "Datos para PPTO" in wb.sheetnames:
            ws = wb["Datos para PPTO"]
            ws.title = "Datos"
        elif "Datos" not in wb.sheetnames:
            raise ValueError("⚠️ La hoja 'Datos para PPTO' no está en el archivo.")

        # Crear hojas mensuales y adicionales si no existen
        meses = [
            "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
            "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"
        ]
        adicionales = ["Horas Extras", "Resumen Anual"]

        for hoja in adicionales + meses:
            if hoja not in wb.sheetnames:
                wb.create_sheet(hoja)

        # Guardar como archivo estructurado temporal
        archivo_temporal = ruta_archivo.replace(".xlsx", "_estructurado.xlsx")
        wb.save(archivo_temporal)

        return archivo_temporal

    except Exception as e:
        raise Exception(f"⚠️ Error al estructurar el archivo Excel: {e}")

# Función principal: estructura y prepara las hojas
def procesar_archivo_excel(ruta_archivo):
    try:
        # 1. Crear estructura
        ruta_estructurada = estructurar_hojas_excel(ruta_archivo)

        # 2. Leer hoja 'Datos' y filtrar por centro de costo
        df = pd.read_excel(ruta_estructurada, sheet_name="Datos")
        cc_col = "Trabajo - Centro de Costo"
        rut_col = "Empleado - Número de Documento"
        fecha_col = "Trabajo - Fecha Ingreso Compañía"

        if cc_col not in df.columns or rut_col not in df.columns or fecha_col not in df.columns:
            raise ValueError("⚠️ Faltan columnas necesarias en la hoja 'Datos'.")

        # Filtrar por centro de costo entre 201 y 501
        df[cc_col] = pd.to_numeric(df[cc_col], errors='coerce')
        df_filtrado = df[df[cc_col].between(201, 501)].copy()

        # Convertir fecha a datetime
        df_filtrado[fecha_col] = pd.to_datetime(df_filtrado[fecha_col], errors='coerce')

        # Ordenar y eliminar duplicados por RUT (quedarse con el más reciente)
        df_ordenado = df_filtrado.sort_values(by=fecha_col, ascending=False)
        df_sin_duplicados = df_ordenado.drop_duplicates(subset=rut_col, keep='first')

        # 3. Sobrescribir la hoja 'Datos' con los datos limpios
        wb = load_workbook(ruta_estructurada)

        if "Datos" in wb.sheetnames:
            del wb["Datos"]
        ws_datos = wb.create_sheet("Datos", 0)
        for row in dataframe_to_rows(df_sin_duplicados, index=False, header=True):
            ws_datos.append(row)

        # Guardar temporalmente para que las funciones usen la nueva hoja 'Datos'
        wb.save(ruta_estructurada)

        # 4. Rellenar las otras hojas con los datos filtrados
        rellenar_hojas_mensuales(ruta_estructurada)
        rellenar_hoja_horas_extras(ruta_estructurada)

        # 5. Rellenar hoja 'Resumen Anual' con formato y estilo
        rellenar_hoja_resumen_anual(ruta_estructurada)  # ← llamado correcto

        # 6. Guardar archivo final con el nombre 'Armado_PPTO.xlsx'
        ruta_final = ruta_estructurada.replace("_estructurado.xlsx", "_Armado_PPTO.xlsx")
        os.rename(ruta_estructurada, ruta_final)

        return ruta_final

    except Exception as e:
        raise Exception(f"⚠️ Error durante el procesamiento del Excel: {e}")
