# relleno_HE.py

import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Alignment, Border, Side

def rellenar_hoja_horas_extras(ruta_archivo_excel):
    try:
        # Leer la hoja 'Datos'
        df_datos = pd.read_excel(ruta_archivo_excel, sheet_name="Datos")

        # Filtrar y seleccionar columnas necesarias
        df_filtrado = df_datos[
            (df_datos["Trabajo - Centro de Costo"] >= 201) &
            (df_datos["Trabajo - Centro de Costo"] <= 501)
        ][[
            "Trabajo - Centro de Costo",
            "Empleado - Número de Documento",
            "Trabajo - Sueldo Base",
            "Campos Personalizados de Trabajo - Jornada"
        ]]

        df_filtrado.sort_values(
            by=["Trabajo - Centro de Costo", "Empleado - Número de Documento"],
            inplace=True
        )

        # Cargar archivo y hoja
        wb = load_workbook(ruta_archivo_excel)
        if "Horas Extras" not in wb.sheetnames:
            raise Exception("La hoja 'Horas Extras' no existe en el archivo Excel.")
        ws: Worksheet = wb["Horas Extras"]
        ws.delete_rows(idx=1, amount=ws.max_row)

        # Encabezados base
        columnas_base = [
            "Trabajo - Centro de Costo",
            "Empleado - Número de Documento",
            "Trabajo - Sueldo Base",
            "Campos Personalizados de Trabajo - Jornada"
        ]

        # Meses y nuevas subcolumnas actualizadas
        meses = [
            "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
            "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"
        ]
        subcolumnas = [
            "HE 55% 6x1", "HE 100% 6x1", "HE 150% 6x1",
            "HE 50% 6x2", "HE 80% 6x2", "HE 100% 6x2",
            "HE 120% 6x2", "HE 50% 5x2"
        ]

        # Fila 1 (meses)
        for col_index, header in enumerate(columnas_base, start=1):
            ws.cell(row=1, column=col_index, value=header)
            ws.merge_cells(start_row=1, start_column=col_index, end_row=2, end_column=col_index)

        col = len(columnas_base) + 1
        for mes in meses:
            ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + len(subcolumnas) - 1)
            ws.cell(row=1, column=col, value=mes)
            for i, subcol in enumerate(subcolumnas):
                ws.cell(row=2, column=col + i, value=subcol)
            col += len(subcolumnas)

        # Centrar texto y resaltar encabezados de meses con borde negro
        borde_negro = Border(
            top=Side(border_style="thin", color="000000"),
            bottom=Side(border_style="thin", color="000000"),
            left=Side(border_style="thin", color="000000"),
            right=Side(border_style="thin", color="000000")
        )

        for row in ws.iter_rows(min_row=1, max_row=2):
            for cell in row:
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = borde_negro

        # Insertar datos con celdas vacías para HE
        for fila in df_filtrado.itertuples(index=False):
            datos_base = list(fila)
            celdas_vacias = [""] * (len(meses) * len(subcolumnas))
            ws.append(datos_base + celdas_vacias)

        wb.save(ruta_archivo_excel)
        print("✅ Hoja 'Horas Extras' actualizada con nuevas subcolumnas ordenadas y destacadas.")

    except Exception as e:
        print(f"⚠️ Error al rellenar la hoja 'Horas Extras': {e}")
